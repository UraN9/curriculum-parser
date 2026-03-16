"""
ETL Service for Academic Curriculum Parsing

Generates a structured Excel workbook from curriculum input data.
Responsibilities:
  - Extracts data from source Excel file with "План" sheet
  - Transforms and aggregates curriculum information by sections and themes
  - Loads formatted output to "Структура.xlsx" with proper styling
  - Logs validation and processing errors for audit trail
  
Key features:
  - Handles duplicate themes across different semesters
  - Calculates totals for hours by type (lectures, practical, lab work, etc.)
  - Applies professional formatting (bold, merged cells, centered alignment, color fills)
  - Automatically adjusts column widths
  - Comprehensive error logging and reporting
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import uuid

from .validation import validate_plan_data, format_validation_report
from .etl_logger import ETLSession, log_validation_error, SEVERITY_ERROR, SEVERITY_WARNING
from .db_loader import (
    load_activity_types,
    load_control_forms,
    find_or_create_semester,
    save_section,
    save_theme,
    save_activity,
    extract_semester_number,
    extract_activity_type,
    extract_control_form,
    commit_changes,
    refresh_summaries
)
from app.database import SessionLocal


# ============================================================================
# Constants
# ============================================================================

BOLD_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
SUMMARY_ROW_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

SECTION_MARKERS = ("РОЗДІЛ",)
THEME_MARKER = "Тема"
ACTIVITY_TYPES = ("Лекція", "Лабораторна", "Практична", "Самостійна")
SUMMARY_KEYWORDS = ("РОЗДІЛ", "РАЗОМ", "ВСЬОГО")

HOUR_COLUMN_TOTAL = 1
HOUR_COLUMN_LECTURES = 3
HOUR_COLUMN_PRACTICAL_LAB = 4
HOUR_COLUMN_SELF_WORK = 5


# ============================================================================
# Data Models
# ============================================================================

def _create_empty_theme() -> dict:
    """Create a dictionary representing an empty theme with hour counters and activities list."""
    return {
        "section": None,
        "theme": None,
        "total": 0,
        "lectures": 0,
        "practical": 0,
        "lab": 0,
        "individual": 0,
        "self": 0,
        "activities": []  # List of individual activities for database storage
    }


# ============================================================================
# ETL Processing Functions
# ============================================================================

def _extract_and_aggregate_data(input_file: str) -> tuple:
    """
    Extract curriculum data from Excel file and aggregate by sections and themes.
    
    Args:
        input_file: Path to input Excel file with "План" sheet
        
    Returns:
        Tuple of:
        - sections: List of unique section names in order
        - themes: Dict mapping (section, theme_name) to aggregated hour data (includes activities list)
        - grand_totals: Dict with global hour statistics
        - semester_number: Extracted semester number from file (e.g., 5 from "5 СЕМЕСТР")
    """
    # Load plan sheet without headers
    df_plan = pd.read_excel(input_file, sheet_name="План", header=None)
    
    # Initialize aggregation structures
    sections = []
    themes = {}
    grand_totals = {
        "total": 0,
        "lectures": 0,
        "practical": 0,
        "lab": 0,
        "individual": 0,
        "self": 0
    }
    
    current_section = None
    current_theme = None
    semester_number = None  # Will be extracted from "X СЕМЕСТР" row
    
    # Column index for control form (Column G = index 6)
    CONTROL_FORM_COLUMN = 6
    
    # Parse each row in the plan sheet
    for _, row in df_plan.iterrows():
        label = str(row[0]).strip() if pd.notnull(row[0]) else ""
        
        # Detect semester row (e.g., "5 СЕМЕСТР")
        if "СЕМЕСТР" in label and semester_number is None:
            semester_number = extract_semester_number(label)
            continue
        
        # Detect section header
        if label.startswith("РОЗДІЛ"):
            current_section = label
            sections.append(current_section)
            current_theme = None  # Reset theme when moving to new section
            continue
        
        # Detect theme header
        if label.startswith("Тема"):
            current_theme = label
            key = (current_section, current_theme)
            if key not in themes:
                theme_data = _create_empty_theme()
                theme_data["section"] = current_section
                theme_data["theme"] = current_theme
                themes[key] = theme_data
            continue
        
        # Process activity rows (Лекція, Лабораторна, etc.)
        if current_theme and any(label.startswith(activity) for activity in ACTIVITY_TYPES):
            # Extract hours from columns with safe defaults
            total_hours = row[HOUR_COLUMN_TOTAL] if pd.notnull(row[HOUR_COLUMN_TOTAL]) else 0
            lectures = row[HOUR_COLUMN_LECTURES] if pd.notnull(row[HOUR_COLUMN_LECTURES]) else 0
            prac_lab_hours = row[HOUR_COLUMN_PRACTICAL_LAB] if pd.notnull(row[HOUR_COLUMN_PRACTICAL_LAB]) else 0
            self_work_hours = row[HOUR_COLUMN_SELF_WORK] if pd.notnull(row[HOUR_COLUMN_SELF_WORK]) else 0
            
            # Extract control form from column G
            control_form_value = row[CONTROL_FORM_COLUMN] if len(row) > CONTROL_FORM_COLUMN else None
            control_form_id = extract_control_form(control_form_value)
            
            key = (current_section, current_theme)
            theme_data = themes[key]
            
            # Determine activity hours based on type
            if label.startswith("Лекція"):
                activity_hours = int(lectures) if lectures else 0
                theme_data["lectures"] += lectures
            elif label.startswith(("Практична", "Семінарська")):
                activity_hours = int(prac_lab_hours) if prac_lab_hours else 0
                theme_data["practical"] += prac_lab_hours
            elif label.startswith("Лабораторна"):
                activity_hours = int(prac_lab_hours) if prac_lab_hours else 0
                theme_data["lab"] += prac_lab_hours
            elif label.startswith("Самостійна"):
                activity_hours = int(self_work_hours) if self_work_hours else 0
                theme_data["self"] += self_work_hours
            else:
                activity_hours = 0
            
            # Store individual activity for database
            activity_info = {
                "name": label,
                "type_id": extract_activity_type(label),
                "hours": activity_hours,
                "control_form_id": control_form_id
            }
            theme_data["activities"].append(activity_info)
            
            # Calculate row total and aggregate
            row_total = total_hours or (lectures + prac_lab_hours + self_work_hours)
            theme_data["total"] += row_total
            
            # Update global totals
            grand_totals["total"] += row_total
            grand_totals["lectures"] += lectures
            grand_totals["practical"] += prac_lab_hours if label.startswith("Практична") else 0
            grand_totals["lab"] += prac_lab_hours if label.startswith("Лабораторна") else 0
            grand_totals["self"] += self_work_hours
    
    # Default semester if not found
    if semester_number is None:
        semester_number = 5
    
    return sections, themes, grand_totals, semester_number


def _build_structure_table(sections: list, themes: dict, grand_totals: dict) -> list:
    """
    Build the structure table with header rows, section content, and summary rows.
    
    Args:
        sections: List of section names
        themes: Dict of aggregated theme data
        grand_totals: Dict with global hour statistics
        
    Returns:
        List of lists representing the structured table
    """
    # Header rows (rows 1-4 in Excel)
    structure_data = [
        ["Назви змістових модулів і тем", "Кількість годин", "", "", "", "", ""],
        ["", "денна форма", "", "", "", "", ""],
        ["", "усього", "у тому числі", "", "", "", ""],
        ["", "", "лекції", "практичні, семінарські", "лабораторні", 
         "індивідуальні завдання", "самостійна робота"]
    ]
    
    # Content rows: sections and themes
    for section in sections:
        # Add section header row
        structure_data.append([section, "", "", "", "", "", ""])
        
        # Find all themes in this section
        section_themes = [t for t in themes.values() if t["section"] == section]
        
        # Add theme rows
        for theme in section_themes:
            structure_data.append([
                theme["theme"],
                theme["total"],
                theme["lectures"],
                theme["practical"],
                theme["lab"],
                theme["individual"],
                theme["self"]
            ])
        
        # Add section summary row
        total_hours = sum(t["total"] for t in section_themes)
        total_lectures = sum(t["lectures"] for t in section_themes)
        total_practical = sum(t["practical"] for t in section_themes)
        total_lab = sum(t["lab"] for t in section_themes)
        total_individual = sum(t["individual"] for t in section_themes)
        total_self = sum(t["self"] for t in section_themes)
        
        section_num = section.split()[1].rstrip('.')
        structure_data.append([
            f"Разом за розділом {section_num}",
            total_hours,
            total_lectures,
            total_practical,
            total_lab,
            total_individual,
            total_self
        ])
    
    # Add grand total row
    structure_data.append([
        "ВСЬОГО ПО НАВЧАЛЬНІЙ ДИСЦИПЛІНІ:",
        grand_totals["total"],
        grand_totals["lectures"],
        grand_totals["practical"],
        grand_totals["lab"],
        grand_totals["individual"],
        grand_totals["self"]
    ])
    
    return structure_data


def _write_data_to_worksheet(ws, structure_data: list) -> list:
    """
    Write raw data to worksheet and return row indices of section headers.
    
    Args:
        ws: Openpyxl worksheet object
        structure_data: List of lists representing table rows
        
    Returns:
        List of row indices (1-based) for section header rows
    """
    section_row_indices = []
    
    for row_idx, row_data in enumerate(structure_data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    # Identify section header rows (those with section names)
    for row_idx in range(5, ws.max_row + 1):  # Skip header rows 1-4
        cell_value = str(ws.cell(row=row_idx, column=1).value or "").strip()
        if cell_value.startswith("РОЗДІЛ"):
            section_row_indices.append(row_idx)
    
    return section_row_indices


def _apply_header_formatting(ws):
    """Apply formatting to header rows (rows 1-4)."""
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = BOLD_FONT
            cell.alignment = CENTER_ALIGNMENT


def _apply_content_formatting(ws):
    """Apply formatting to data rows (rows 5 and below)."""
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        first_cell = row[0]
        
        # Bold formatting for section/summary rows
        if first_cell.value:
            cell_text = str(first_cell.value).upper()
            if any(keyword in cell_text for keyword in SUMMARY_KEYWORDS):
                first_cell.font = BOLD_FONT
        
        # Center alignment for numeric columns (B-G)
        for cell in row[1:7]:
            if cell.value is not None:
                cell.alignment = CENTER_ALIGNMENT


def _apply_summary_row_styling(ws):
    """Apply background color to summary rows."""
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        first_cell = row[0]
        cell_text = str(first_cell.value or "").strip().upper()
        
        # Highlight summary rows with gray background
        if cell_text.startswith("РАЗОМ ЗА РОЗДІЛОМ") or cell_text.startswith("ВСЬОГО ПО НАВЧАЛЬНІЙ"):
            first_cell.font = BOLD_FONT
            for cell in row[:7]:  # Columns A-G
                cell.fill = SUMMARY_ROW_FILL


def _merge_header_cells(ws):
    """Merge header cells for proper table structure."""
    # First column merged for header label
    ws.merge_cells("A1:A4")
    
    # Hour category headers
    ws.merge_cells("B1:G1")  # "Кількість годин"
    ws.merge_cells("B2:G2")  # "денна форма"
    ws.merge_cells("C3:G3")  # "у тому числі"
    ws.merge_cells("B3:B4")  # "усього"


def _merge_section_cells(ws, section_row_indices: list):
    """Merge cells in section header rows across all columns."""
    for row_idx in section_row_indices:
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        
        # Apply formatting to merged cell
        cell = ws.cell(row=row_idx, column=1)
        cell.alignment = CENTER_ALIGNMENT
        cell.font = BOLD_FONT


def _auto_adjust_column_widths(ws):
    """Automatically adjust column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_content_length = 0
        
        # Find maximum content length in this column
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    content_length = len(str(cell.value))
                    max_content_length = max(max_content_length, content_length)
            except Exception:
                # Skip merged cells and other problematic cells
                continue
        
        # Apply width with padding (more width for column A)
        width_padding = 4 if col_idx == 1 else 2.5
        ws.column_dimensions[col_letter].width = max_content_length + width_padding


# ============================================================================
# Main Function
# ============================================================================

def generate_structure(
    input_file: str, 
    output_file: str = "Структура.xlsx",
    discipline_id: int = 1,
    save_to_database: bool = True
) -> None:
    """
    Main ETL function: Extract → Transform → Load (Excel + Database).
    
    Processes curriculum data from input Excel file and generates
    a properly formatted structure workbook. Optionally saves data
    to PostgreSQL database. Includes comprehensive validation and error logging.
    
    Args:
        input_file: Path to input Excel file containing "План" sheet
        output_file: Path where output Excel file will be saved
        discipline_id: ID of the discipline in database (default: 1)
        save_to_database: Whether to save data to PostgreSQL (default: True)
        
    Raises:
        FileNotFoundError: If input file does not exist
        ValueError: If "План" sheet not found in input file or validation fails
    """
    # Create ETL session for this run
    etl_session = ETLSession(file_name=input_file)
    print(f"\n📊 Starting ETL process: {etl_session}")
    
    # Load the plan sheet
    try:
        df_plan = pd.read_excel(input_file, sheet_name="План", header=None)
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {input_file}")
    except ValueError as e:
        raise ValueError(f"Sheet 'План' not found in {input_file}")
    
    # === VALIDATE ===
    print("\n✓ Validating data...")
    validation_result = validate_plan_data(df_plan)
    
    # Print validation report
    print(format_validation_report(validation_result))
    
    # Log validation errors/warnings to database
    if validation_result.errors or validation_result.warnings:
        db_session = SessionLocal()
        try:
            # Log all errors
            for issue in validation_result.errors:
                log_validation_error(
                    db_session=db_session,
                    message=issue.message,
                    row_number=issue.row_number,
                    field_name=issue.column,
                    source_data=issue.issue_type,
                    etl_session_id=etl_session.session_id,
                    file_name=input_file,
                    severity=SEVERITY_ERROR
                )
            # Log all warnings
            for issue in validation_result.warnings:
                log_validation_error(
                    db_session=db_session,
                    message=issue.message,
                    row_number=issue.row_number,
                    field_name=issue.column,
                    source_data=issue.issue_type,
                    etl_session_id=etl_session.session_id,
                    file_name=input_file,
                    severity=SEVERITY_WARNING
                )
            db_session.commit()
            print(f"  ✓ Logged {len(validation_result.errors)} errors and {len(validation_result.warnings)} warnings to etl_errors")
        except Exception as e:
            db_session.rollback()
            print(f"  ⚠ Failed to log validation issues to database: {e}")
        finally:
            db_session.close()
    
    # Stop if critical errors found
    if not validation_result.is_valid:
        raise ValueError(
            f"Validation failed with {validation_result.error_count} critical error(s). "
            "Please review and correct the input file."
        )
    
    # === EXTRACT ===
    print("✓ Extracting and aggregating data...")
    sections, themes, grand_totals, semester_number = _extract_and_aggregate_data(input_file)
    
    # === TRANSFORM ===
    print("✓ Transforming data...")
    structure_data = _build_structure_table(sections, themes, grand_totals)
    
    # === LOAD TO EXCEL ===
    print("✓ Loading and formatting Excel output...")
    # Create workbook and worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Структура"
    
    # Write data and get section row positions
    section_row_indices = _write_data_to_worksheet(worksheet, structure_data)
    
    # Apply formatting
    _merge_header_cells(worksheet)
    _merge_section_cells(worksheet, section_row_indices)
    _apply_header_formatting(worksheet)
    _apply_content_formatting(worksheet)
    _apply_summary_row_styling(worksheet)
    _auto_adjust_column_widths(worksheet)
    
    # Save workbook
    workbook.save(output_file)
    print(f"  ✓ Excel file saved: {output_file}")
    
    # === LOAD TO DATABASE ===
    db_stats = {"sections": 0, "themes": 0, "activities": 0}
    
    if save_to_database:
        print("✓ Loading data to database...")
        db_session = SessionLocal()
        
        try:
            # Load reference data (idempotent - safe to call multiple times)
            load_activity_types(db_session)
            load_control_forms(db_session)
            
            # Find or create semester
            semester_id = find_or_create_semester(db_session, semester_number)
            
            # Process sections → themes → activities
            for section_name in sections:
                # Save section
                section_id = save_section(
                    db_session,
                    name=section_name,
                    discipline_id=discipline_id,
                    semester_id=semester_id
                )
                db_stats["sections"] += 1
                
                # Find all themes in this section
                section_themes = [
                    theme_data for key, theme_data in themes.items()
                    if theme_data["section"] == section_name
                ]
                
                for theme_data in section_themes:
                    # Save theme
                    theme_id = save_theme(
                        db_session,
                        name=theme_data["theme"],
                        section_id=section_id,
                        total_hours=int(theme_data["total"])
                    )
                    db_stats["themes"] += 1
                    
                    # Save individual activities
                    for activity in theme_data["activities"]:
                        save_activity(
                            db_session,
                            name=activity["name"],
                            type_id=activity["type_id"],
                            hours=activity["hours"],
                            theme_id=theme_id,
                            control_form_id=activity["control_form_id"]
                        )
                        db_stats["activities"] += 1
            
            # Commit all changes
            commit_changes(db_session)
            print(f"  ✓ Database load completed: {db_stats['sections']} sections, "
                  f"{db_stats['themes']} themes, {db_stats['activities']} activities")
            
            # === REFRESH SUMMARY VIEWS ===
            print("✓ Refreshing summary views...")
            refresh_result = refresh_summaries(db_session)
            if refresh_result["success"]:
                print(f"  ✓ Refreshed {refresh_result['views_refreshed']} materialized views")
            else:
                print(f"  ⚠ Summary refresh skipped: {refresh_result['error']}")
            
        except Exception as e:
            db_session.rollback()
            print(f"  ✗ Database load failed: {e}")
            raise
        finally:
            db_session.close()
    else:
        print("  ⏭ Database load skipped (save_to_database=False)")
    
    # Success summary
    print(f"\n{'='*70}")
    print(f"✓ ETL PROCESS COMPLETED SUCCESSFULLY")
    print(f"{'='*70}")
    print(f"  Output file:    {output_file}")
    print(f"  Semester:       {semester_number}")
    print(f"  Sections:       {len(sections)}")
    print(f"  Themes:         {len(themes)}")
    print(f"  Total hours:    {grand_totals['total']}")
    if save_to_database:
        print(f"  DB Activities:  {db_stats['activities']}")
    print(f"  Session ID:     {etl_session.session_id}")
    print(f"{'='*70}\n")


# ============================================================================
# Entry Point
# ============================================================================

def run_etl_pipeline(
    input_file: str,
    discipline_id: int = 1,
    output_file: str = None,
    idempotent: bool = True
) -> dict:
    """
    Run ETL pipeline with idempotent support for async execution.
    
    This is the main entry point for async ETL tasks. It wraps
    generate_structure() and returns statistics for tracking.
    
    Args:
        input_file: Path to input Excel file
        discipline_id: ID of the discipline in database
        output_file: Optional output file path (default: Структура.xlsx)
        idempotent: If True, use UPSERT logic (default: True)
        
    Returns:
        dict: Result with statistics:
            - records_processed: Total records processed
            - records_created: New records created
            - records_updated: Existing records updated
            - records_skipped: Records skipped (validation errors)
            - summary: Dict with detailed counts by entity type
    """
    if output_file is None:
        output_file = "Структура.xlsx"
    
    # Track statistics
    stats = {
        'records_processed': 0,
        'records_created': 0,
        'records_updated': 0,
        'records_skipped': 0,
        'summary': {}
    }
    
    # Create ETL session for this run
    etl_session = ETLSession(file_name=input_file)
    print(f"\n📊 Starting ETL pipeline (idempotent={idempotent}): {etl_session}")
    
    # Load the plan sheet
    try:
        df_plan = pd.read_excel(input_file, sheet_name="План", header=None)
    except FileNotFoundError:
        raise FileNotFoundError(f"Input file not found: {input_file}")
    except ValueError as e:
        raise ValueError(f"Sheet 'План' not found in {input_file}")
    
    # === VALIDATE ===
    print("\n✓ Validating data...")
    validation_result = validate_plan_data(df_plan)
    print(format_validation_report(validation_result))
    
    # Log validation errors/warnings
    db_session = SessionLocal()
    try:
        if validation_result.errors or validation_result.warnings:
            for issue in validation_result.errors:
                log_validation_error(
                    db_session=db_session,
                    message=issue.message,
                    row_number=issue.row_number,
                    field_name=issue.column,
                    source_data=issue.issue_type,
                    etl_session_id=etl_session.session_id,
                    file_name=input_file,
                    severity=SEVERITY_ERROR
                )
            for issue in validation_result.warnings:
                log_validation_error(
                    db_session=db_session,
                    message=issue.message,
                    row_number=issue.row_number,
                    field_name=issue.column,
                    source_data=issue.issue_type,
                    etl_session_id=etl_session.session_id,
                    file_name=input_file,
                    severity=SEVERITY_WARNING
                )
            db_session.commit()
            stats['records_skipped'] = len(validation_result.errors)
        
        if not validation_result.is_valid:
            raise ValueError(
                f"Validation failed with {validation_result.error_count} critical error(s)."
            )
        
        # === EXTRACT ===
        print("✓ Extracting and aggregating data...")
        sections, themes, grand_totals, semester_number = _extract_and_aggregate_data(input_file)
        
        # === TRANSFORM ===
        print("✓ Transforming data...")
        structure_data = _build_structure_table(sections, themes, grand_totals)
        
        # === LOAD TO EXCEL ===
        print("✓ Loading and formatting Excel output...")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Структура"
        
        section_row_indices = _write_data_to_worksheet(worksheet, structure_data)
        _merge_header_cells(worksheet)
        _merge_section_cells(worksheet, section_row_indices)
        _apply_header_formatting(worksheet)
        _apply_content_formatting(worksheet)
        _apply_summary_row_styling(worksheet)
        _auto_adjust_column_widths(worksheet)
        workbook.save(output_file)
        print(f"  ✓ Excel file saved: {output_file}")
        
        # === LOAD TO DATABASE (IDEMPOTENT) ===
        print("✓ Loading data to database...")
        
        # Load reference data (idempotent)
        load_activity_types(db_session)
        load_control_forms(db_session)
        
        # Track created/updated
        created_sections = 0
        created_themes = 0
        created_activities = 0
        updated_themes = 0
        updated_activities = 0
        
        # Find or create semester
        semester_id = find_or_create_semester(db_session, semester_number)
        
        # Process sections → themes → activities
        for section_name in sections:
            # Check if section exists
            from app.models import Section, Theme, Activity
            existing_section = db_session.query(Section).filter_by(
                name=section_name,
                discipline_id=discipline_id,
                semester_id=semester_id
            ).first()
            
            section_id = save_section(
                db_session,
                name=section_name,
                discipline_id=discipline_id,
                semester_id=semester_id
            )
            
            if not existing_section:
                created_sections += 1
            
            # Find all themes in this section
            section_themes = [
                theme_data for key, theme_data in themes.items()
                if theme_data["section"] == section_name
            ]
            
            for theme_data in section_themes:
                # Check if theme exists
                existing_theme = db_session.query(Theme).filter_by(
                    name=theme_data["theme"],
                    section_id=section_id
                ).first()
                
                theme_id = save_theme(
                    db_session,
                    name=theme_data["theme"],
                    section_id=section_id,
                    total_hours=int(theme_data["total"])
                )
                
                if existing_theme:
                    updated_themes += 1
                else:
                    created_themes += 1
                
                # Save individual activities
                for activity in theme_data["activities"]:
                    existing_activity = db_session.query(Activity).filter_by(
                        name=activity["name"],
                        theme_id=theme_id
                    ).first()
                    
                    save_activity(
                        db_session,
                        name=activity["name"],
                        type_id=activity["type_id"],
                        hours=activity["hours"],
                        theme_id=theme_id,
                        control_form_id=activity["control_form_id"]
                    )
                    
                    if existing_activity:
                        updated_activities += 1
                    else:
                        created_activities += 1
        
        # Commit all changes
        commit_changes(db_session)
        
        # Update stats
        stats['records_processed'] = created_sections + created_themes + created_activities + updated_themes + updated_activities
        stats['records_created'] = created_sections + created_themes + created_activities
        stats['records_updated'] = updated_themes + updated_activities
        stats['summary'] = {
            'sections': {'created': created_sections},
            'themes': {'created': created_themes, 'updated': updated_themes},
            'activities': {'created': created_activities, 'updated': updated_activities},
            'semester': semester_number
        }
        
        print(f"  ✓ Database load completed:")
        print(f"    - Sections: {created_sections} created")
        print(f"    - Themes: {created_themes} created, {updated_themes} updated")
        print(f"    - Activities: {created_activities} created, {updated_activities} updated")
        
        # Refresh summary views
        print("✓ Refreshing summary views...")
        refresh_result = refresh_summaries(db_session)
        if refresh_result["success"]:
            print(f"  ✓ Refreshed {refresh_result['views_refreshed']} materialized views")
        
        print(f"\n{'='*70}")
        print(f"✓ ETL PIPELINE COMPLETED SUCCESSFULLY")
        print(f"{'='*70}\n")
        
        return stats
        
    except Exception as e:
        db_session.rollback()
        raise
    finally:
        db_session.close()


if __name__ == "__main__":
    generate_structure("НПр КН 2025.xlsx")
